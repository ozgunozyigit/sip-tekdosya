import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import re
import math
import os
import calendar
from io import BytesIO
from datetime import datetime, date, timedelta, timezone
from rapidfuzz import fuzz


# =========================================================
# SAYFA AYARI
# =========================================================

st.set_page_config(
    page_title="Eczane Sipariş Motoru",
    layout="wide"
)


# =========================================================
# SABİT LİSTE DIŞI BAŞLANGIÇLAR
# =========================================================

LISTE_DISI_BASLANGICLAR_RAW = [
    "CUBITAN",
    "DIASIP",
    "ENSURE",
    "FORTIMEL",
    "FORTINI",
    "GLUCERNA",
    "IMPACT",
    "NUTRISON",
    "NUTRIVIGOR",
    "PEDIASURE",
    "PEPTAMEN",
    "RESOURCE",
    "ENJEKTOR",
    "MAJISTRAL",
    "SENTE",
    "SARIINTRAKET",
    "INFATRINI",
    "VITAL",
    "SANITAYARA",
]


# =========================================================
# TÜRKÇE TARİH / RAPOR ARALIĞI
# =========================================================

TURKCE_AYLAR = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}


def bugun_turkiye():
    return (datetime.now(timezone.utc) + timedelta(hours=3)).date()


def ay_ekle(tarih, ay_sayisi):
    yeni_ay_index = tarih.month - 1 + ay_sayisi
    yeni_yil = tarih.year + yeni_ay_index // 12
    yeni_ay = yeni_ay_index % 12 + 1
    yeni_gun = min(tarih.day, calendar.monthrange(yeni_yil, yeni_ay)[1])
    return date(yeni_yil, yeni_ay, yeni_gun)


def turkce_tarih_yaz(tarih):
    return f"{tarih.day:02d} {TURKCE_AYLAR[tarih.month]} {tarih.year}"


def onerilen_rapor_araligi():
    bugun = bugun_turkiye()
    bu_ayin_ilk_gunu = date(bugun.year, bugun.month, 1)
    baslangic = ay_ekle(bu_ayin_ilk_gunu, -3)
    bitis = bu_ayin_ilk_gunu - timedelta(days=1)
    return baslangic, bitis


# =========================================================
# RESMİ TATİL / İŞ GÜNÜ HESABI
# =========================================================

def turkiye_resmi_tatilleri(yil):
    tatiller = set()

    tatiller.add(date(yil, 1, 1))
    tatiller.add(date(yil, 4, 23))
    tatiller.add(date(yil, 5, 1))
    tatiller.add(date(yil, 5, 19))
    tatiller.add(date(yil, 7, 15))
    tatiller.add(date(yil, 8, 30))
    tatiller.add(date(yil, 10, 29))

    dini_tatiller = {
        2025: [
            date(2025, 3, 30), date(2025, 3, 31), date(2025, 4, 1),
            date(2025, 6, 6), date(2025, 6, 7), date(2025, 6, 8), date(2025, 6, 9),
        ],
        2026: [
            date(2026, 3, 20), date(2026, 3, 21), date(2026, 3, 22),
            date(2026, 5, 27), date(2026, 5, 28), date(2026, 5, 29), date(2026, 5, 30),
        ],
        2027: [
            date(2027, 3, 9), date(2027, 3, 10), date(2027, 3, 11),
            date(2027, 5, 16), date(2027, 5, 17), date(2027, 5, 18), date(2027, 5, 19),
        ],
        2028: [
            date(2028, 2, 26), date(2028, 2, 27), date(2028, 2, 28),
            date(2028, 5, 4), date(2028, 5, 5), date(2028, 5, 6), date(2028, 5, 7),
        ],
        2029: [
            date(2029, 2, 14), date(2029, 2, 15), date(2029, 2, 16),
            date(2029, 4, 23), date(2029, 4, 24), date(2029, 4, 25), date(2029, 4, 26),
        ],
        2030: [
            date(2030, 2, 4), date(2030, 2, 5), date(2030, 2, 6),
            date(2030, 4, 13), date(2030, 4, 14), date(2030, 4, 15), date(2030, 4, 16),
        ],
    }

    for tatil in dini_tatiller.get(yil, []):
        tatiller.add(tatil)

    return tatiller


def is_gunu_mu(tarih):
    if tarih.weekday() >= 5:
        return False

    if tarih in turkiye_resmi_tatilleri(tarih.year):
        return False

    return True


def ay_is_gunu_bilgisi(referans_tarih):
    ay_ilk_gun = date(referans_tarih.year, referans_tarih.month, 1)
    ay_son_gun = date(
        referans_tarih.year,
        referans_tarih.month,
        calendar.monthrange(referans_tarih.year, referans_tarih.month)[1]
    )

    toplam_is_gunu = 0
    kalan_is_gunu = 0

    gun = ay_ilk_gun

    while gun <= ay_son_gun:
        if is_gunu_mu(gun):
            toplam_is_gunu += 1

            if gun >= referans_tarih:
                kalan_is_gunu += 1

        gun += timedelta(days=1)

    toplam_is_gunu = max(toplam_is_gunu, 1)
    kalan_is_gunu = max(kalan_is_gunu, 1)

    return toplam_is_gunu, kalan_is_gunu, ay_son_gun


# =========================================================
# ÜRÜN ADI NORMALİZE
# =========================================================

def normalize_urun_adi(text):
    if pd.isna(text):
        return ""

    text = str(text)

    # ÜBS tarafındaki küçük l / büyük I karışıklığı için
    text = text.replace("l", "I")

    text = text.upper()

    replacements = {
        "İ": "I",
        "İ": "I",
        "ı": "I",
        "Ğ": "G",
        "Ü": "U",
        "Ş": "S",
        "Ö": "O",
        "Ç": "C",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    # % işareti bilerek kaldırılmıyor.
    for ch in [".", ",", ";", ":", "/", "\\", "-", "_", "(", ")", "[", "]", "*"]:
        text = text.replace(ch, "")

    text = re.sub(r"\s+", "", text)

    # Yüzde sonrası O/0 karışıklığı: %O15 -> %015
    text = text.replace("%O", "%0")

    # Sayıdan sonra gelen O harfini 0 yapar: 2OO -> 200
    for _ in range(5):
        text = re.sub(r"(\d)O", r"\g<1>0", text)

    return text


LISTE_DISI_BASLANGICLAR = [
    normalize_urun_adi(x) for x in LISTE_DISI_BASLANGICLAR_RAW
]


def liste_disi_baslangic_bul(normalize_key):
    normalize_key = str(normalize_key)

    for raw, key in zip(LISTE_DISI_BASLANGICLAR_RAW, LISTE_DISI_BASLANGICLAR):
        if normalize_key.startswith(key):
            return True, raw

    return False, ""


def gorunen_urun_adi_olustur(normalize_key):
    text = str(normalize_key)

    if text == "" or text == "NAN":
        return ""

    # %015 gibi ifadeleri %0.15 şeklinde gösterir.
    # Örnek: ALPHAGANP%015GOZ -> ALPHAGANP %0.15 GOZ
    text = re.sub(r"%0(\d{2})(?=[A-Z])", r" %0.\1 ", text)

    # %5DEXTROZ gibi ifadeleri %5 DEXTROZ şeklinde gösterir.
    text = re.sub(r"%(\d+(?:\.\d+)?)(?=[A-Z])", r" %\1 ", text)

    # Tek başına G kaldırıldı. Çünkü 15GOZ ifadesinde 15 G OZ hatası oluşturuyordu.
    birimler = [
        "MCG", "MG", "ML", "CM", "MM", "CC", "IU", "GR"
    ]

    birim_regex = "|".join(sorted(birimler, key=len, reverse=True))

    text = re.sub(
        rf"(\d+)({birim_regex})(?=[A-Z]|$)",
        r"\1 \2 ",
        text
    )

    text = re.sub(r"(?<=[A-Z])(?=\d)", " ", text)
    text = re.sub(r"(?<=\d)(?=[A-Z])", " ", text)

    kelimeler = [
        "TABLET",
        "KAPSUL",
        "KAPSEL",
        "KAPLET",
        "FILM",
        "FLAKON",
        "AMPUL",
        "SASE",
        "SASET",
        "DAMLA",
        "SURUP",
        "SUSPANSIYON",
        "SUSP",
        "SOLUSYON",
        "KREM",
        "MERHEM",
        "JEL",
        "SPREY",
        "PASTIL",
        "PED",
        "SAFT",
        "KASE",
        "ADET",
        "ORAL",
        "NAZAL",
        "GOZ",
        "KULAK",
        "BURUN",
        "DERI",
        "CILT",
        "UZATILMIS",
        "SALIMLI",
        "ENTERIK",
        "KAPLI",
        "RETARD",
        "FORT",
        "FORTE",
        "PLUS",
        "PEDIATRIK",
        "YETISKIN",
        "COCUK",
        "BEBEK",
        "AROMALI",
        "MUZ",
        "CILEK",
        "PORTAKAL",
        "ORMAN",
        "MEYVE",
        "ENJEKTOR",
        "INSULIN",
        "SUPER",
        "KOMPRES",
        "GAZ",
        "STERIL",
        "ENERJI",
        "ENERGY",
        "COZELTI",
    ]

    kelimeler = sorted(set(kelimeler), key=len, reverse=True)

    for kelime in kelimeler:
        text = re.sub(rf"(?<!^)(?<!\s)({kelime})", r" \1", text)

    text = re.sub(r"\s+", " ", text).strip()

    return text


# =========================================================
# SAYI OKUMA / FORMATLAMA
# =========================================================

def to_number(value):
    if pd.isna(value):
        return 0

    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return 0
        except Exception:
            pass
        return float(value)

    value = str(value).strip()

    if value == "":
        return 0

    value = value.replace(",", ".")

    try:
        return float(value)
    except Exception:
        return 0


def sayi_formatla(value):
    if isinstance(value, (int, float)):
        value = round(float(value), 2)

        if value.is_integer():
            return str(int(value))

        return f"{value:.2f}".rstrip("0").rstrip(".")

    return value


# =========================================================
# MASTER DOSYA
# urun_master.xlsx
# Sadece isim standardizasyonu için kullanılır.
# B kolonu: standart ürün adı
# C kolonu kullanılmaz.
# =========================================================

def master_dosya_yolu_bul():
    try:
        app_klasoru = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        app_klasoru = os.getcwd()

    return os.path.join(app_klasoru, "urun_master.xlsx")


@st.cache_data(show_spinner=False)
def oku_master_standart_isimler_cached(master_path, master_mtime):
    if not os.path.exists(master_path):
        return {}, pd.DataFrame(), False, {
            "master_okundu": False,
            "master_toplam_satir": 0,
            "master_standart_key_sayisi": 0,
        }

    try:
        master_df = pd.read_excel(master_path, usecols="B", header=None)
    except Exception as e:
        raise ValueError(
            "Master dosya okunamadı. Lütfen urun_master.xlsx dosyasını kontrol edin."
        ) from e

    master_df.columns = ["standart_urun_adi"]
    master_toplam_satir = len(master_df)

    master_df["standart_urun_adi"] = master_df["standart_urun_adi"].astype(str).str.strip()
    master_df["master_key"] = master_df["standart_urun_adi"].apply(normalize_urun_adi)

    baslik_keyleri = {
        "URUNADI",
        "URUN",
        "PRODUCTNAME",
        "ILACADI",
        "ADI",
        "MASTERURUNADI",
        "STANDARTURUNADI",
    }

    master_df = master_df[
        (master_df["master_key"] != "")
        & (master_df["master_key"] != "NAN")
        & (~master_df["master_key"].isin(baslik_keyleri))
    ].copy()

    master_df = master_df.drop_duplicates(subset=["master_key"], keep="first")

    standart_isim_map = dict(
        zip(master_df["master_key"], master_df["standart_urun_adi"])
    )

    debug = {
        "master_okundu": True,
        "master_toplam_satir": master_toplam_satir,
        "master_standart_key_sayisi": len(standart_isim_map),
    }

    return standart_isim_map, master_df, True, debug


def oku_master_standart_isimler():
    master_path = master_dosya_yolu_bul()

    if not os.path.exists(master_path):
        return {}, pd.DataFrame(), False, {
            "master_okundu": False,
            "master_toplam_satir": 0,
            "master_standart_key_sayisi": 0,
        }

    master_mtime = os.path.getmtime(master_path)

    return oku_master_standart_isimler_cached(master_path, master_mtime)


# =========================================================
# ÜBS DOSYASI OKUMA
# A: Ürün adı
# B: 3 aylık toplam satış
# F: Stok
# =========================================================

@st.cache_data(show_spinner=False)
def oku_ubs_tek_dosya_cached(file_bytes):
    try:
        df = pd.read_excel(BytesIO(file_bytes), usecols="A,B,F")
    except Exception as e:
        raise ValueError(
            "Dosya okunamadı. Lütfen Eczanem programından alınan "
            "Ürün Bazında Satış raporunu Excel formatında yükleyin."
        ) from e

    df.columns = ["urun_adi", "toplam_3ay_satis", "stok"]

    df["normalize_ad"] = df["urun_adi"].apply(normalize_urun_adi)
    df["toplam_3ay_satis"] = df["toplam_3ay_satis"].apply(to_number)
    df["stok"] = df["stok"].apply(to_number)

    df = df[df["normalize_ad"] != ""]
    df = df[df["normalize_ad"] != "NAN"]

    # Aynı ürün farklı fiyatlardan satılmış olabilir.
    # Satışlar toplanır; stok güncel değer olduğu için toplanmaz, ilk değer alınır.
    sonuc = (
        df.groupby("normalize_ad", as_index=False)
        .agg(
            toplam_3ay_satis=("toplam_3ay_satis", "sum"),
            stok=("stok", "first"),
            ornek_ubs_adi=("urun_adi", "first")
        )
    )

    return sonuc


# =========================================================
# SİPARİŞ DURUMU
# =========================================================

def siparis_durumu_belirle(row):
    ortalama_satis = row["ortalama_satis"]
    ham_siparis = row["ham_siparis_miktari"]
    hesap_stok = row["hesap_stok"]

    if (
        ortalama_satis > 10
        and ham_siparis > 0
        and hesap_stok < (ortalama_satis / 30) * 7
        and hesap_stok <= 30
    ):
        return "ACİL"

    if ham_siparis > 0:
        return "SİPARİŞ"

    return "GEREK YOK"


def siparis_onceligi_belirle(durum):
    priority_map = {
        "ACİL": 1,
        "SİPARİŞ": 2,
        "GEREK YOK": 3
    }

    return priority_map.get(durum, 99)


# =========================================================
# SİPARİŞ HESAPLAMA
# =========================================================

def siparis_hesapla(file_bytes):
    sonuc = oku_ubs_tek_dosya_cached(file_bytes)

    bugun = bugun_turkiye()
    toplam_is_gunu, kalan_is_gunu, ay_son_gun = ay_is_gunu_bilgisi(bugun)

    standart_isim_map, master_df, master_var_mi, master_debug = oku_master_standart_isimler()

    sonuc["gorunen_urun_adi"] = sonuc["normalize_ad"].map(standart_isim_map)
    sonuc["gorunen_urun_adi"] = sonuc["gorunen_urun_adi"].fillna(
        sonuc["normalize_ad"].apply(gorunen_urun_adi_olustur)
    )

    liste_disi_sonuclari = sonuc["normalize_ad"].apply(liste_disi_baslangic_bul)
    sonuc["liste_disi"] = liste_disi_sonuclari.apply(lambda x: x[0])
    sonuc["liste_disi_sebebi"] = liste_disi_sonuclari.apply(
        lambda x: f"Başlangıç: {x[1]}" if x[0] else ""
    )

    haric_tutulan_ubs = sonuc[sonuc["liste_disi"] == True].copy()
    sonuc = sonuc[sonuc["liste_disi"] == False].copy()

    if not haric_tutulan_ubs.empty:
        haric_tutulan_ubs = haric_tutulan_ubs.rename(columns={
            "gorunen_urun_adi": "Ürün Adı",
            "toplam_3ay_satis": "3 Aylık Satış",
            "stok": "Stok",
            "liste_disi_sebebi": "Liste Dışı Sebebi",
        })

        haric_tutulan_ubs = haric_tutulan_ubs[
            [
                "Ürün Adı",
                "3 Aylık Satış",
                "Stok",
                "Liste Dışı Sebebi",
            ]
        ]
    else:
        haric_tutulan_ubs = pd.DataFrame(
            columns=[
                "Ürün Adı",
                "3 Aylık Satış",
                "Stok",
                "Liste Dışı Sebebi",
            ]
        )

    sonuc["ortalama_satis"] = sonuc["toplam_3ay_satis"] / 3
    sonuc["ortalama_satis"] = sonuc["ortalama_satis"].round(2)

    sonuc["hesap_stok"] = sonuc["stok"].apply(lambda x: max(x, 0))

    sonuc["ortalama_gunluk_satis"] = sonuc["ortalama_satis"] / toplam_is_gunu
    sonuc["ortalama_gunluk_satis"] = sonuc["ortalama_gunluk_satis"].round(4)

    sonuc["kalan_ay_ihtiyaci"] = sonuc["ortalama_gunluk_satis"] * kalan_is_gunu
    sonuc["kalan_ay_ihtiyaci"] = sonuc["kalan_ay_ihtiyaci"].round(2)

    sonuc["ham_siparis_miktari"] = sonuc["kalan_ay_ihtiyaci"] - sonuc["hesap_stok"]
    sonuc["ham_siparis_miktari"] = sonuc["ham_siparis_miktari"].apply(
        lambda x: max(0, math.ceil(x))
    )

    def parca_sayisi_belirle(ham_siparis):
        if ham_siparis > 300:
            return 7
        elif ham_siparis > 250:
            return 6
        elif ham_siparis > 200:
            return 5
        elif ham_siparis > 150:
            return 4
        elif ham_siparis > 100:
            return 3
        elif ham_siparis > 60:
            return 2
        else:
            return 1

    sonuc["parca_sayisi"] = sonuc["ham_siparis_miktari"].apply(parca_sayisi_belirle)

    sonuc["planlanan_siparis_miktari"] = sonuc.apply(
        lambda row: math.ceil(row["ham_siparis_miktari"] / row["parca_sayisi"])
        if row["ham_siparis_miktari"] > 0 else 0,
        axis=1
    )

    sonuc["siparis_durumu"] = sonuc.apply(siparis_durumu_belirle, axis=1)
    sonuc["siparis_onceligi"] = sonuc["siparis_durumu"].apply(siparis_onceligi_belirle)

    sayisal_kolonlar = [
        "planlanan_siparis_miktari",
        "ham_siparis_miktari",
        "parca_sayisi",
        "toplam_3ay_satis",
        "ortalama_satis",
        "ortalama_gunluk_satis",
        "kalan_ay_ihtiyaci",
        "stok",
        "hesap_stok",
    ]

    for kolon in sayisal_kolonlar:
        sonuc[kolon] = sonuc[kolon].round(2)

    sonuc = sonuc.sort_values(
        by=["siparis_onceligi", "ham_siparis_miktari", "ortalama_satis"],
        ascending=[True, False, False]
    ).reset_index(drop=True)

    sonuc = sonuc[
        [
            "gorunen_urun_adi",
            "planlanan_siparis_miktari",
            "ham_siparis_miktari",
            "parca_sayisi",
            "toplam_3ay_satis",
            "ortalama_satis",
            "stok",
            "siparis_durumu",
        ]
    ]

    sonuc = sonuc.rename(columns={
        "gorunen_urun_adi": "Ürün Adı",
        "planlanan_siparis_miktari": "Parti Sipariş Mik.",
        "ham_siparis_miktari": "Top. Sipariş Mik.",
        "parca_sayisi": "Parti Sayısı",
        "toplam_3ay_satis": "3 Aylık Satış",
        "ortalama_satis": "Ort. Aylık Satış",
        "stok": "Stok",
        "siparis_durumu": "Durum",
    })

    is_gunu_bilgi = {
        "bugun": bugun,
        "toplam_is_gunu": toplam_is_gunu,
        "kalan_is_gunu": kalan_is_gunu,
        "ay_son_gun": ay_son_gun,
    }

    debug_bilgi = {
        **master_debug,
        "liste_disi_baslangic_sayisi": len(LISTE_DISI_BASLANGICLAR_RAW),
        "ubs_normalize_urun_sayisi": len(sonuc) + len(haric_tutulan_ubs),
        "ubs_liste_disi_cikarilan_sayisi": len(haric_tutulan_ubs),
    }

    return sonuc, haric_tutulan_ubs, master_df, master_var_mi, is_gunu_bilgi, debug_bilgi


# =========================================================
# EXCEL FORMAT / İNDİRME
# =========================================================

def excel_sayfa_formatla(worksheet):
    from openpyxl.styles import Font, Alignment

    max_row = worksheet.max_row

    headers = {}
    for cell in worksheet[1]:
        headers[cell.column] = cell.value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            header = headers.get(cell.column)

            if header == "Ürün Adı":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if isinstance(cell.value, (int, float)):
                value = round(float(cell.value), 2)

                if value.is_integer():
                    cell.value = int(value)
                    cell.number_format = "0"
                else:
                    cell.value = value
                    cell.number_format = "0.00"

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 45)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.freeze_panes = "A2"


def excel_indir(df, haric_df=None):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sipariş Sonuç")

        if haric_df is not None:
            haric_df.to_excel(writer, index=False, sheet_name="Listeden Çıkarılanlar")

        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            excel_sayfa_formatla(worksheet)

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            if max_row < 1 or max_col < 1:
                continue

            last_col_letter = get_column_letter(max_col)
            table_range = f"A1:{last_col_letter}{max_row}"

            safe_table_name = (
                sheet_name.replace(" ", "")
                .replace("ı", "i")
                .replace("İ", "I")
                .replace("ç", "c")
                .replace("Ç", "C")
                .replace("ğ", "g")
                .replace("Ğ", "G")
                .replace("ö", "o")
                .replace("Ö", "O")
                .replace("ş", "s")
                .replace("Ş", "S")
                .replace("ü", "u")
                .replace("Ü", "U")
            )

            table = Table(displayName=f"Tablo_{safe_table_name}", ref=table_range)

            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False
            )

            table.tableStyleInfo = style
            worksheet.add_table(table)
            worksheet.auto_filter.ref = table_range

        worksheet = writer.sheets["Sipariş Sonuç"]

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        durum_col_index = None
        for cell in worksheet[1]:
            if cell.value == "Durum":
                durum_col_index = cell.column
                break

        acil_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        gerek_yok_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

        if durum_col_index is not None:
            for row_num in range(2, max_row + 1):
                durum = worksheet.cell(row=row_num, column=durum_col_index).value

                if durum == "ACİL":
                    fill = acil_fill
                elif durum == "GEREK YOK":
                    fill = gerek_yok_fill
                else:
                    fill = None

                if fill:
                    for col_num in range(1, max_col + 1):
                        worksheet.cell(row=row_num, column=col_num).fill = fill

    return output.getvalue()


# =========================================================
# PDF İNDİRME
# =========================================================

def pdf_indir(df):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        from xml.sax.saxutils import escape
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "PDF oluşturmak için reportlab paketi kurulu olmalıdır. "
            "CMD üzerinden şu komutu çalıştırın: py -m pip install reportlab"
        ) from e

    output = BytesIO()

    def font_bul():
        font_paths = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "DejaVuSans.ttf",
        ]

        for path in font_paths:
            if os.path.exists(path):
                return path

        return None

    font_path = font_bul()

    if font_path:
        try:
            pdfmetrics.registerFont(TTFont("TRFont", font_path))
            font_name = "TRFont"
        except Exception:
            font_name = "Helvetica"
    else:
        font_name = "Helvetica"

    pdf = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleTR",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=15,
        leading=18,
        alignment=1,
        spaceAfter=8
    )

    normal_style = ParagraphStyle(
        "NormalTR",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7.2,
        leading=8.6
    )

    header_style = ParagraphStyle(
        "HeaderTR",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7.2,
        leading=8.6,
        alignment=1
    )

    elements = []

    pdf_df = df[df["Durum"] == "ACİL"].copy()

    pdf_df = pdf_df[
        [
            "Ürün Adı",
            "Parti Sipariş Mik.",
            "Top. Sipariş Mik.",
            "Ort. Aylık Satış",
            "Stok",
        ]
    ]

    toplam_urun = len(pdf_df)
    tarih = datetime.now().strftime("%d.%m.%Y %H:%M")

    elements.append(Paragraph("ACİL SİPARİŞ LİSTESİ", title_style))
    elements.append(
        Paragraph(
            f"Tarih: {tarih} &nbsp;&nbsp; | &nbsp;&nbsp; "
            f"Acil Sipariş Ürün Sayısı: {toplam_urun}",
            normal_style
        )
    )
    elements.append(Spacer(1, 8))

    def sayi_pdf_formatla(value):
        try:
            value = round(float(value), 2)
            if value.is_integer():
                return str(int(value))
            return f"{value:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(value)

    table_data = [
        [
            Paragraph("Ürün Adı", header_style),
            Paragraph("Parti Sip. Mik.", header_style),
            Paragraph("Top. Sip. Mik.", header_style),
            Paragraph("Ort. Aylık", header_style),
            Paragraph("Stok", header_style),
        ]
    ]

    for _, row in pdf_df.iterrows():
        table_data.append(
            [
                Paragraph(escape(str(row["Ürün Adı"])), normal_style),
                sayi_pdf_formatla(row["Parti Sipariş Mik."]),
                sayi_pdf_formatla(row["Top. Sipariş Mik."]),
                sayi_pdf_formatla(row["Ort. Aylık Satış"]),
                sayi_pdf_formatla(row["Stok"]),
            ]
        )

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            8.7 * cm,
            2.7 * cm,
            2.7 * cm,
            2.4 * cm,
            1.7 * cm,
        ]
    )

    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 7.2),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    elements.append(table)

    def sayfa_numarasi(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.drawRightString(20.0 * cm, 0.5 * cm, f"Sayfa {doc.page}")
        canvas.restoreState()

    pdf.build(elements, onFirstPage=sayfa_numarasi, onLaterPages=sayfa_numarasi)

    return output.getvalue()


# =========================================================
# TABLO ARAMA / RENKLENDİRME / SCROLL
# =========================================================

def tablo_ara(sonuc, arama):
    if not arama:
        filtre = sonuc.copy()
        filtre["Arama Skoru"] = ""
        return filtre

    arama_key = normalize_urun_adi(arama)

    sonuc_goster = sonuc.copy()

    def arama_skoru(row):
        urun_adi_gorunen = str(row["Ürün Adı"])
        urun_key = normalize_urun_adi(urun_adi_gorunen)

        if len(arama_key) >= 3:
            if urun_key.startswith(arama_key):
                return 100

        if arama_key in urun_key:
            return 95

        skor1 = fuzz.partial_ratio(urun_key, arama_key)
        skor2 = fuzz.token_set_ratio(urun_key, arama_key)

        return max(skor1, skor2)

    sonuc_goster["Arama Skoru"] = sonuc_goster.apply(arama_skoru, axis=1)

    filtre = sonuc_goster[sonuc_goster["Arama Skoru"] >= 55].copy()
    filtre = filtre.sort_values(
        ["Arama Skoru", "Top. Sipariş Mik."],
        ascending=[False, False]
    )

    return filtre


def satir_renklendir(row):
    if row.get("Durum") == "ACİL":
        return ["background-color: #FDECEC;"] * len(row)
    return [""] * len(row)


def scroll_to_results():
    components.html(
        """
        <script>
            const doc = window.parent.document;
            const target = doc.getElementById("sonuc-bolumu");
            if (target) {
                target.scrollIntoView({ behavior: "smooth", block: "start" });
            }
        </script>
        """,
        height=0,
    )


# =========================================================
# WEB ARAYÜZ
# =========================================================

st.title("Eczane Sipariş Motoru")
st.caption("Tek Excel dosyasından akıllı sipariş önerisi")

bugun = bugun_turkiye()
toplam_is_gunu, kalan_is_gunu, ay_son_gun = ay_is_gunu_bilgisi(bugun)
aktif_ay_adi = TURKCE_AYLAR[bugun.month]

rapor_baslangic, rapor_bitis = onerilen_rapor_araligi()
rapor_araligi_metni = f"{turkce_tarih_yaz(rapor_baslangic)} - {turkce_tarih_yaz(rapor_bitis)}"

with st.expander("📌 Kullanım kılavuzunu görüntüle", expanded=False):
    st.markdown("### ✅ ÜBS Satış Raporu")

    st.info(
        f"📅 Önerilen Rapor Tarih Aralığı: **{rapor_araligi_metni}**"
    )

    st.write(
        "Eczanem programından geçmiş 3 tamamlanmış ayı kapsayan "
        "“Ürün Bazında Satış” raporunu tek Excel dosyası olarak kaydedin."
    )

    st.write(
        "Raporu sipariş vereceğiniz gün oluşturun. "
        "Dosyanın içindeki **Stok Mik.** kolonu güncel stok olarak kullanılacaktır."
    )

    st.warning(
        "Dosya yapısı: **A sütunu Ürün Adı**, "
        "**B sütunu 3 aylık toplam satış**, "
        "**F sütunu Stok Mik.**"
    )


col1, col2 = st.columns([1, 1])

with col1:
    st.subheader("Ürün Bazında Satış Raporu")

    with st.container(border=True):
        ubs_file = st.file_uploader(
            "Belirtilen tarihlere göre alınmış 3 aylık Ürün Bazında Satış Excel dosyasını yükleyin",
            type=["xls", "xlsx"]
        )

        siparis_hazirla_clicked = st.button(
            "Sipariş Listesini Oluştur",
            type="primary",
            use_container_width=True,
            disabled=not bool(ubs_file)
        )

with col2:
    st.subheader("Rapor Bilgisi")

    with st.container(border=True):
        st.caption("Önerilen rapor aralığı")
        st.markdown(f"**{rapor_araligi_metni}**")

        st.caption("Bugün")
        st.markdown(f"**{turkce_tarih_yaz(bugun)}**")

        st.caption(f"{aktif_ay_adi} ayı toplam resmi iş günü")
        st.markdown(f"**{toplam_is_gunu}**")

        st.info(
            f"⏳ **{aktif_ay_adi} ayının sonuna kadar kalan resmi iş günü: {kalan_is_gunu}**"
        )

        st.caption(
            "Sipariş hesabı kalan resmi iş günü ihtiyacına göre yapılır."
        )

        st.caption(
            "Resmi tatil hesabı otomatik yapılır. "
            "Sonradan ilan edilen idari izinler tekrar kontrol edilmelidir."
        )

st.divider()

if not ubs_file:
    st.info("Sipariş hazırlamak için lütfen 3 aylık ÜBS satış dosyasını yükleyin.")

else:
    if siparis_hazirla_clicked:
        try:
            with st.spinner("Sipariş listesi hazırlanıyor..."):
                file_bytes = ubs_file.getvalue()

                (
                    sonuc,
                    haric_tutulan_ubs,
                    master_df,
                    master_var_mi,
                    is_gunu_bilgi,
                    debug_bilgi,
                ) = siparis_hesapla(file_bytes)

                st.session_state["sonuc"] = sonuc
                st.session_state["haric_tutulan_ubs"] = haric_tutulan_ubs
                st.session_state["master_df"] = master_df
                st.session_state["master_var_mi"] = master_var_mi
                st.session_state["is_gunu_bilgi"] = is_gunu_bilgi
                st.session_state["debug_bilgi"] = debug_bilgi
                st.session_state["scroll_to_results"] = True

            st.success("Hesaplama tamamlandı.")

        except Exception as e:
            st.error(f"Hesaplama sırasında hata oluştu: {e}")


if "sonuc" in st.session_state:
    st.markdown('<div id="sonuc-bolumu"></div>', unsafe_allow_html=True)

    sonuc = st.session_state["sonuc"]
    haric_tutulan_ubs = st.session_state.get("haric_tutulan_ubs", pd.DataFrame())
    master_var_mi = st.session_state.get("master_var_mi", False)
    is_gunu_bilgi = st.session_state.get("is_gunu_bilgi", {})
    debug_bilgi = st.session_state.get("debug_bilgi", {})

    toplam_acil = len(sonuc[sonuc["Durum"] == "ACİL"])
    toplam_siparis = len(sonuc[sonuc["Durum"] == "SİPARİŞ"])
    toplam_gerek_yok = len(sonuc[sonuc["Durum"] == "GEREK YOK"])
    toplam_haric = len(haric_tutulan_ubs)

    kpi1, kpi2, kpi3, kpi4 = st.columns(4)

    kpi1.metric("ACİL", toplam_acil)
    kpi2.metric("SİPARİŞ", toplam_siparis)
    kpi3.metric("GEREK YOK", toplam_gerek_yok)
    kpi4.metric("LİSTE DIŞI", toplam_haric)

    kalan_is_gunu_hesap = is_gunu_bilgi.get("kalan_is_gunu", kalan_is_gunu)

    st.info(
        f"Sipariş önerileri, ay sonuna kadar kalan **{kalan_is_gunu_hesap} resmi iş günü** "
        f"ihtiyacına göre hesaplanmıştır."
    )

    if master_var_mi:
        st.info(
            "Master dosya okundu. Ürün isimleri mümkün olduğu ölçüde master listedeki standart isimlerle gösteriliyor."
        )
    else:
        st.warning(
            "Master dosya bulunamadı. Ürün isimleri normalize edilmiş ÜBS adına göre gösteriliyor. "
            "Liste dışı başlangıç kuralları yine çalışır."
        )

    if toplam_haric > 0:
        st.info(
            f"Başlangıç kuralına göre {toplam_haric} ürün sipariş listesinden çıkarıldı."
        )
    else:
        st.info("Başlangıç kuralına göre listeden çıkarılan ürün bulunamadı.")

    with st.expander("Kontrol / Debug Bilgisi", expanded=False):
        st.write(f"Master dosya okundu mu: **{debug_bilgi.get('master_okundu', False)}**")
        st.write(f"Master toplam satır: **{debug_bilgi.get('master_toplam_satir', 0)}**")
        st.write(f"Master standart key sayısı: **{debug_bilgi.get('master_standart_key_sayisi', 0)}**")
        st.write(f"Liste dışı başlangıç sayısı: **{debug_bilgi.get('liste_disi_baslangic_sayisi', 0)}**")
        st.write(f"ÜBS normalize ürün sayısı: **{debug_bilgi.get('ubs_normalize_urun_sayisi', 0)}**")
        st.write(
            f"ÜBS içinde başlangıç kuralıyla çıkarılan ürün sayısı: "
            f"**{debug_bilgi.get('ubs_liste_disi_cikarilan_sayisi', 0)}**"
        )

        if not haric_tutulan_ubs.empty:
            st.markdown("#### İlk 20 çıkarılan ürün")
            st.dataframe(
                haric_tutulan_ubs.head(20),
                use_container_width=True,
                hide_index=True
            )

    st.divider()

    download_col1, download_col2, download_col3 = st.columns([1, 1, 2])

    with download_col1:
        try:
            excel_data = excel_indir(
                st.session_state["sonuc"],
                st.session_state.get("haric_tutulan_ubs"),
            )

            st.download_button(
                label="Excel İndir",
                data=excel_data,
                file_name="siparis_sonuc.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Excel oluşturulamadı: {e}")

    with download_col2:
        try:
            pdf_data = pdf_indir(st.session_state["sonuc"])

            st.download_button(
                label="Acil Sipariş PDF İndir",
                data=pdf_data,
                file_name="acil_siparis_listesi.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"PDF oluşturulamadı: {e}")

    arama = st.text_input("Ürün adı ara")

    filtre = tablo_ara(sonuc, arama)
    filtre_gorunum = filtre.drop(columns=["Arama Skoru"], errors="ignore")

    st.subheader("Sipariş Listesi")

    st.dataframe(
        filtre_gorunum.style
        .apply(satir_renklendir, axis=1)
        .format(sayi_formatla),
        use_container_width=True,
        hide_index=True
    )

    with st.expander("Sipariş listesinden çıkarılan ürünler", expanded=True):
        if haric_tutulan_ubs.empty:
            st.info("Sipariş listesinden çıkarılan ürün yok.")
        else:
            st.dataframe(
                haric_tutulan_ubs.style.format(sayi_formatla),
                use_container_width=True,
                hide_index=True
            )

    if st.session_state.get("scroll_to_results"):
        scroll_to_results()
        st.session_state["scroll_to_results"] = False
